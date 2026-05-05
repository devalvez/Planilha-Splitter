import pandas as pd
import os
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class ExcelWriter:
    @staticmethod
    def save(df, output_path, format='.xlsx', smart_format=False, **kwargs):
        extension = format.lower()
        if not extension.startswith('.'):
            extension = '.' + extension
            
        final_path = output_path
        if not output_path.lower().endswith(extension):
            final_path += extension

        try:
            if extension == '.csv':
                sep = kwargs.get('sep', ',')
                encoding = kwargs.get('encoding', 'utf-8')
                index = kwargs.get('index', False)
                df.to_csv(final_path, sep=sep, encoding=encoding, index=index)
            elif extension == '.ods':
                index = kwargs.get('index', False)
                df.to_excel(final_path, engine='odf', index=index)
            elif smart_format and extension == '.xlsx':
                ExcelWriter._save_with_smart_format(df, final_path, **kwargs)
            else: # .xlsx padrão
                index = kwargs.get('index', False)
                df.to_excel(final_path, index=index)
            return True
        except Exception as e:
            print(f"Erro ao salvar arquivo {final_path}: {e}")
            import traceback
            traceback.print_exc()
            return False

    @staticmethod
    def _save_with_smart_format(df, path, **kwargs):
        """
        Salva o DataFrame mesclando colunas com base no maior conteúdo encontrado.
        Limite de 5 colunas para mesclagem.
        """
        index = kwargs.get('index', False)
        
        # 1. Encontrar o maior número de caracteres em qualquer célula do DF
        if df.empty:
            max_chars = 0
        else:
            # Converte para string e calcula o tamanho de cada célula
            # Usamos fillna('') e str(x) para garantir que nenhum float/NaN chegue ao len()
            max_chars = df.fillna('').astype(str).map(len).max().max()
            
            # Caso o max() ainda retorne algo não numérico (raro)
            if not isinstance(max_chars, (int, float)) or math.isnan(max_chars):
                max_chars = 0
        
        # 2. Calcular quantas colunas são necessárias (Assumindo aprox. 20 caracteres por coluna padrão)
        # Se max_chars for 100, precisaria de 5 colunas.
        chars_per_col = 20
        needed_cols = min(math.ceil(max_chars / chars_per_col), 5)
        if needed_cols < 1: needed_cols = 1

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=index, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Se needed_cols > 1, precisamos re-estruturar a planilha
            # NOTA: O pandas já escreveu os dados. Para mesclar sem perder dados, 
            # o ideal seria escrever os dados pulando colunas, mas o mais simples
            # e visualmente eficaz para o usuário é ajustar a largura e mesclar 
            # as colunas adjacentes se elas estiverem vazias ou se quisermos expandir.
            
            # Como o usuário quer que TODAS as colunas sigam o padrão:
            # Vamos aplicar a mesclagem em blocos de 'needed_cols'.
            
            rows = len(df) + 1 # +1 para o cabeçalho
            cols = len(df.columns)
            if index: cols += 1

            # Configuração de alinhamento para quebra de linha
            alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

            if needed_cols > 1:
                # Limpamos o sheet e reescrevemos manualmente para garantir o espaçamento correto
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.value = None
                
                # Reescrever dados com espaçamento
                # Cabeçalho
                headers = []
                if index: headers.append(df.index.name or '')
                headers.extend(df.columns)
                
                for c_idx, header in enumerate(headers):
                    start_col = (c_idx * needed_cols) + 1
                    end_col = start_col + needed_cols - 1
                    cell = worksheet.cell(row=1, column=start_col)
                    cell.value = header
                    cell.alignment = alignment
                    if needed_cols > 1:
                        worksheet.merge_cells(
                            start_row=1, start_column=start_col, 
                            end_row=1, end_column=end_col
                        )

                # Dados
                for r_idx, row_data in enumerate(df.values):
                    current_row = r_idx + 2
                    row_list = []
                    if index: row_list.append(df.index[r_idx])
                    row_list.extend(row_data)
                    
                    for c_idx, value in enumerate(row_list):
                        start_col = (c_idx * needed_cols) + 1
                        end_col = start_col + needed_cols - 1
                        cell = worksheet.cell(row=current_row, column=start_col)
                        cell.value = value
                        cell.alignment = alignment
                        if needed_cols > 1:
                            worksheet.merge_cells(
                                start_row=current_row, start_column=start_col, 
                                end_row=current_row, end_column=end_col
                            )
            else:
                # Apenas uma coluna, aplicar wrap_text
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = alignment

            # Ajustar largura das colunas para "fit"
            for i in range(1, (cols * needed_cols) + 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 15
